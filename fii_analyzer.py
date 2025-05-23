import os

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def verificar_arquivo_existente(nome_arquivo):
    """Verifica se o arquivo já existe e pergunta ao usuário se deseja substituí-lo"""
    if os.path.exists(nome_arquivo):
        print(f"\nAVISO: O arquivo '{nome_arquivo}' já existe no diretório.")
        # substituir = input("Deseja substituí-lo? (S/N): ").strip().upper()
        # return substituir == 'S'
    return True


def scrape_fundamentus_fiis():
    url = "https://www.fundamentus.com.br/fii_resultado.php"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table', {'id': 'tabelaResultado'})

        if not table:
            print("Erro: Não foi possível encontrar a tabela de dados no site.")
            return None

        # Extrair cabeçalhos
        headers = [th.get_text(strip=True) for th in table.find('tr').find_all('th')]

        # Extrair dados das linhas
        data = []
        for row in table.find_all('tr')[1:]:
            row_data = [td.get_text(strip=True) for td in row.find_all('td')]
            data.append(row_data)

        # Criar DataFrame
        df = pd.DataFrame(data, columns=headers)

        # Renomear colunas para consistência
        df = df.rename(columns={
            'Papel': 'Papel',
            'Segmento': 'Segmento',
            'Cotação': 'Cotacao',
            'FFO Yield': 'FFO Yield',
            'Dividend Yield': 'Dividend Yield',
            'P/VP': 'P/VP',
            'Valor de Mercado': 'Valor de Mercado',
            'Liquidez': 'Liquidez',
            'Qtd de imóveis': 'Qtd Imoveis',
            'Preço do m2': 'Preco m2',
            'Aluguel por m2': 'Aluguel m2',
            'Cap Rate': 'Cap Rate',
            'Vacância Média': 'Vacancia Media'
        })

        return df

    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar o site: {e}")
        return None
    except Exception as e:
        print(f"Erro inesperado: {e}")
        return None


def clean_and_convert_data(df):
    try:
        # Converter valores numéricos
        df['Dividend Yield'] = df['Dividend Yield'].str.replace('%', '').str.replace('.', '').str.replace(',',
                                                                                                          '.').astype(
            float) / 100
        df['P/VP'] = df['P/VP'].str.replace(',', '.').astype(float)
        df['Valor de Mercado'] = df['Valor de Mercado'].apply(
            lambda x: float(x.replace('.', '')) if isinstance(x, str) else x)
        df['Liquidez'] = df['Liquidez'].apply(lambda x: float(x.replace('.', '')) if isinstance(x, str) else x)
        df['Vacancia Media'] = df['Vacancia Media'].str.replace('%', '').str.replace(',', '.').astype(float) / 100

        return df
    except Exception as e:
        print(f"Erro ao converter dados: {e}")
        return None


def apply_filters(df):
    try:
        # Aplicar filtros
        filtered_df = df[
            (df['Dividend Yield'] > 0.7) &
            (df['Dividend Yield'] < 0.25) &
            (df['P/VP'] > 0.5) &
            (df['P/VP'] < 1.1) &
            (df['Liquidez'] > 1000000) &
            (df['Valor de Mercado'] > 1000000000)
            ].copy()

        return filtered_df
    except Exception as e:
        print(f"Erro ao aplicar filtros: {e}")
        return None


def calculate_score(row):
    try:
        score = 0

        # Dividend Yield (quanto maior melhor)
        if row['Dividend Yield'] >= 0.14:
            score += 2
        elif row['Dividend Yield'] >= 0.12:
            score += 1

        # P/VP (quanto menor melhor)
        if row['P/VP'] <= 0.80:
            score += 2
        elif row['P/VP'] <= 0.85:
            score += 1

        # Liquidez (quanto maior melhor)
        if row['Liquidez'] >= 5000000:
            score += 2
        elif row['Liquidez'] >= 2000000:
            score += 1

        # Valor de Mercado (quanto maior melhor)
        if row['Valor de Mercado'] >= 2000000000:
            score += 2
        elif row['Valor de Mercado'] >= 1500000000:
            score += 1

        # Vacância Média (quanto menor melhor)
        if row['Vacancia Media'] <= 0.05:  # 5%
            score += 2
        elif row['Vacancia Media'] <= 0.10:  # 10%
            score += 1

        return min(score, 10)  # Nota máxima de 10 pontos
    except Exception as e:
        print(f"Erro ao calcular score: {e}")
        return 0


def get_top_by_segment(df, n=5):
    """Retorna os top N FIIs de cada segmento"""
    if df.empty:
        return pd.DataFrame()

    # Agrupar por segmento e pegar os top N de cada grupo
    top_by_segment = df.sort_values(['Segmento', 'Nota', 'Dividend Yield'],
                                    ascending=[True, False, False]) \
        .groupby('Segmento') \
        .head(n) \
        .sort_values(['Segmento', 'Nota'], ascending=[True, False])

    return top_by_segment


def style_excel_output(writer, df, sheet_name='Top por Segmento'):
    try:
        workbook = writer.book
        worksheet = workbook[sheet_name]

        # Definir estilos
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Aplicar estilos ao cabeçalho
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

            # Ajustar largura das colunas
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 17

        # Aplicar estilos aos dados
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Formatar porcentagens
        for col_name in ['Dividend Yield', 'Vacancia Media']:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '0.00%'

        # Formatar P/VP com 2 casas decimais
        pvp_col = df.columns.get_loc('P/VP') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=pvp_col)
            cell.number_format = '0.00'

        # Formatar Valor de Mercado como número com separadores
        vm_col = df.columns.get_loc('Valor de Mercado') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=vm_col)
            cell.number_format = '#,##0'

        # Formatar Qtd de Imóveis como número inteiro
        qtd_col = df.columns.get_loc('Qtd Imoveis') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=qtd_col)
            cell.number_format = '0'

        # Adicionar formatação condicional para a coluna Nota
        nota_col = df.columns.get_loc('Nota') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=nota_col)
            if cell.value >= 8:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif cell.value >= 5:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Congelar painel
        worksheet.freeze_panes = 'A2'
    except Exception as e:
        print(f"Erro ao formatar arquivo Excel: {e}")


def main():
    print("\n=== Filtro de Fundos Imobiliários ===")
    print("Fonte: Fundamentus (https://www.fundamentus.com.br/fii_resultado.php)")

    output_file = "fundos_imobiliarios_filtrados.xlsx"

    # Verificar se o arquivo já existe
    if not verificar_arquivo_existente(output_file):
        print("Operação cancelada pelo usuário.")
        return

    print("\nIniciando raspagem de dados do Fundamentus...")
    df = scrape_fundamentus_fiis()

    if df is not None:
        print("Dados obtidos com sucesso! Processando...")
        df = clean_and_convert_data(df)

        if df is not None:
            filtered_df = apply_filters(df)

            if filtered_df is not None:
                if len(filtered_df) > 0:
                    # Calcular nota para cada fundo
                    filtered_df['Nota'] = filtered_df.apply(calculate_score, axis=1)

                    # Selecionar e ordenar colunas
                    final_df = filtered_df[['Papel', 'Segmento', 'Dividend Yield', 'P/VP',
                                            'Valor de Mercado', 'Liquidez', 'Qtd Imoveis',
                                            'Vacancia Media', 'Nota']]

                    # Ordenar por nota
                    final_df = final_df.sort_values(by=['Nota', 'Dividend Yield'], ascending=[False, False])

                    # Pegar os top 5 de cada segmento
                    top_by_segment = get_top_by_segment(final_df, 5)

                    try:
                        # Salvar em Excel
                        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                            # Salvar todos os fundos filtrados
                            final_df.to_excel(writer, index=False, sheet_name='Todos Fundos')

                            # Salvar os top 5 por segmento
                            if not top_by_segment.empty:
                                top_by_segment.to_excel(writer, index=False, sheet_name='Top por Segmento')
                                style_excel_output(writer, top_by_segment, 'Top por Segmento')

                            # Formatar também a aba de todos os fundos
                            style_excel_output(writer, final_df, 'Todos Fundos')

                        print(f"\nProcesso concluído com sucesso!")
                        print(f"Arquivo salvo como: {os.path.abspath(output_file)}")
                        print(f"Total de fundos encontrados: {len(final_df)}")

                        if not top_by_segment.empty:
                            print("\nTop 5 fundos por segmento:")
                            for segment, group in top_by_segment.groupby('Segmento'):
                                print(f"\nSegmento: {segment}")
                                print(group[['Papel', 'Nota', 'Dividend Yield', 'P/VP']].to_string(index=False))

                    except Exception as e:
                        print(f"\nErro ao salvar o arquivo: {e}")
                else:
                    print("\nNenhum fundo atende aos critérios de filtro especificados.")
            else:
                print("\nErro ao filtrar os dados.")
        else:
            print("\nErro ao processar os dados.")
    else:
        print("\nNão foi possível obter os dados do site.")


if __name__ == "__main__":
    main()